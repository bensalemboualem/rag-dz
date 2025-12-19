from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
from datetime import datetime
from pathlib import Path

class KPIDashboard:
    def generate(self) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "KPIs"
        
        ws.merge_cells("A1:C1")
        ws["A1"] = "IA FACTORY - KPIs"
        ws["A1"].font = Font(size=16, bold=True)
        
        ws["A2"], ws["B2"] = "CH:", "contact@iafactory.ch"
        ws["A3"], ws["B3"] = "DZ:", "contact@iafactoryalgeria.com"
        
        for i, (m, v, t) in enumerate([("MRR", "8,500 CHF", "+15%"), ("Clients", "12", "+3"), ("Margin", "92%", "+2%")], 5):
            ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = m, v, t
        
        for i, (m, r) in enumerate([("Jan", 3000), ("Feb", 4500), ("Mar", 6000), ("Apr", 7500), ("May", 8500)], 10):
            ws[f"A{i}"], ws[f"B{i}"] = m, r
        
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=10, max_row=14))
        chart.set_categories(Reference(ws, min_col=1, min_row=10, max_row=14))
        ws.add_chart(chart, "D10")
        
        Path("outputs/dashboards").mkdir(parents=True, exist_ok=True)
        f = f"outputs/dashboards/kpi_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        wb.save(f)
        return f
