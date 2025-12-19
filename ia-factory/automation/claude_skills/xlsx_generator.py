"""
IA Factory - Claude Skill: XLSX Generator
Génération de fichiers Excel professionnels
"""

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any, Union
from datetime import datetime
from enum import Enum
import os

router = APIRouter(prefix="/skills/xlsx", tags=["Claude Skills - XLSX"])


class SpreadsheetType(str, Enum):
    INVOICE = "invoice"             # Facture
    QUOTE = "quote"                 # Devis
    BUDGET = "budget"               # Budget
    REPORT = "report"               # Rapport
    TIMESHEET = "timesheet"         # Feuille de temps
    INVENTORY = "inventory"         # Inventaire
    CRM = "crm"                     # Contacts CRM
    PROJECT_TRACKER = "project"     # Suivi projet
    FINANCIAL = "financial"         # États financiers
    DASHBOARD = "dashboard"         # Tableau de bord


class ChartType(str, Enum):
    BAR = "bar"
    LINE = "line"
    PIE = "pie"
    COLUMN = "column"
    AREA = "area"
    SCATTER = "scatter"


class CellFormat(BaseModel):
    """Format de cellule"""
    bold: bool = False
    italic: bool = False
    font_size: int = 11
    font_color: Optional[str] = None
    bg_color: Optional[str] = None
    number_format: Optional[str] = None  # Ex: "#,##0.00", "0%", "dd/mm/yyyy"
    alignment: str = "left"  # left, center, right
    border: bool = False


class SheetData(BaseModel):
    """Données d'une feuille"""
    name: str
    headers: List[str]
    rows: List[List[Any]]
    formulas: Dict[str, str] = Field(default_factory=dict)  # Ex: {"E2": "=C2*D2"}
    column_widths: Dict[str, int] = Field(default_factory=dict)  # Ex: {"A": 20}
    charts: List[Dict[str, Any]] = Field(default_factory=list)


class SpreadsheetRequest(BaseModel):
    """Requête de génération de spreadsheet"""
    spreadsheet_type: SpreadsheetType
    title: str
    author: str = "IA Factory"
    sheets: List[SheetData] = Field(default_factory=list)
    include_formulas: bool = True
    include_charts: bool = True
    language: str = "fr"
    currency: str = "CHF"


class GeneratedSpreadsheet(BaseModel):
    """Spreadsheet généré"""
    id: str
    filename: str
    filepath: str
    spreadsheet_type: SpreadsheetType
    sheet_count: int
    row_count: int
    created_at: datetime
    download_url: str


class XlsxGenerator:
    """
    Générateur de fichiers Excel
    Utilise openpyxl pour créer des spreadsheets professionnels
    """
    
    # Templates par type
    TEMPLATES = {
        SpreadsheetType.INVOICE: {
            "sheets": [
                {
                    "name": "Facture",
                    "headers": ["Description", "Quantité", "Prix Unitaire", "TVA %", "Total HT", "Total TTC"],
                    "sample_rows": [
                        ["Service de conseil IA", 10, 150, 7.7, "=B2*C2", "=E2*(1+D2/100)"],
                        ["Développement RAG", 5, 200, 7.7, "=B3*C3", "=E3*(1+D3/100)"],
                        ["Formation équipe", 2, 500, 7.7, "=B4*C4", "=E4*(1+D4/100)"]
                    ],
                    "summary_formulas": {
                        "Total HT": "=SUM(E2:E100)",
                        "TVA": "=SUM(F2:F100)-SUM(E2:E100)",
                        "Total TTC": "=SUM(F2:F100)"
                    }
                }
            ]
        },
        SpreadsheetType.QUOTE: {
            "sheets": [
                {
                    "name": "Devis",
                    "headers": ["Poste", "Description", "Jours", "TJM", "Total"],
                    "sample_rows": [
                        ["Phase 1", "Analyse et conception", 5, 1200, "=C2*D2"],
                        ["Phase 2", "Développement", 15, 1200, "=C3*D3"],
                        ["Phase 3", "Tests et déploiement", 5, 1200, "=C4*D4"],
                        ["Phase 4", "Formation", 2, 1000, "=C5*D5"]
                    ]
                }
            ]
        },
        SpreadsheetType.BUDGET: {
            "sheets": [
                {
                    "name": "Budget",
                    "headers": ["Catégorie", "Jan", "Fév", "Mar", "Avr", "Mai", "Jun", "Total"],
                    "sample_rows": [
                        ["Revenus", 10000, 12000, 15000, 18000, 20000, 25000, "=SUM(B2:G2)"],
                        ["Marketing", 1000, 1200, 1500, 1800, 2000, 2500, "=SUM(B3:G3)"],
                        ["Infrastructure", 500, 500, 600, 600, 700, 800, "=SUM(B4:G4)"],
                        ["Salaires", 5000, 5000, 5000, 6000, 6000, 6000, "=SUM(B5:G5)"],
                        ["Marge", "=B2-B3-B4-B5", "=C2-C3-C4-C5", "=D2-D3-D4-D5", "=E2-E3-E4-E5", "=F2-F3-F4-F5", "=G2-G3-G4-G5", "=SUM(B6:G6)"]
                    ]
                }
            ]
        },
        SpreadsheetType.PROJECT_TRACKER: {
            "sheets": [
                {
                    "name": "Suivi Projet",
                    "headers": ["Tâche", "Responsable", "Date Début", "Date Fin", "Statut", "% Avancement", "Notes"],
                    "sample_rows": [
                        ["Analyse besoins", "Boualem", "2025-01-01", "2025-01-15", "Terminé", 100, ""],
                        ["Design architecture", "Dev Team", "2025-01-16", "2025-01-31", "En cours", 75, ""],
                        ["Développement MVP", "Dev Team", "2025-02-01", "2025-03-15", "À faire", 0, ""],
                        ["Tests", "QA", "2025-03-16", "2025-03-31", "À faire", 0, ""],
                        ["Déploiement", "DevOps", "2025-04-01", "2025-04-07", "À faire", 0, ""]
                    ]
                }
            ]
        },
        SpreadsheetType.CRM: {
            "sheets": [
                {
                    "name": "Contacts",
                    "headers": ["Entreprise", "Contact", "Email", "Téléphone", "Statut", "Dernière Interaction", "Valeur Potentielle", "Notes"],
                    "sample_rows": []
                },
                {
                    "name": "Pipeline",
                    "headers": ["Opportunité", "Client", "Valeur", "Probabilité", "Valeur Pondérée", "Étape", "Date Prévue"],
                    "sample_rows": []
                }
            ]
        },
        SpreadsheetType.DASHBOARD: {
            "sheets": [
                {
                    "name": "KPIs",
                    "headers": ["Métrique", "Valeur Actuelle", "Objectif", "% Atteinte", "Tendance"],
                    "sample_rows": [
                        ["Chiffre d'Affaires", 50000, 60000, "=B2/C2*100", "↑"],
                        ["Nombre de Clients", 15, 20, "=B3/C3*100", "↑"],
                        ["NPS", 75, 80, "=B4/C4*100", "→"],
                        ["Taux de Rétention", 92, 95, "=B5/C5*100", "↑"]
                    ]
                }
            ]
        }
    }
    
    # Couleurs IA Factory
    COLORS = {
        "primary": "1F4E79",
        "secondary": "2E75B6",
        "accent": "BDD7EE",
        "header_bg": "1F4E79",
        "header_text": "FFFFFF",
        "row_alt": "F2F8FC",
        "success": "92D050",
        "warning": "FFC000",
        "error": "FF6B6B"
    }
    
    def __init__(self):
        self.output_dir = "outputs/spreadsheets"
        os.makedirs(self.output_dir, exist_ok=True)
        self.generated_files: Dict[str, GeneratedSpreadsheet] = {}
    
    async def generate(self, request: SpreadsheetRequest) -> GeneratedSpreadsheet:
        """Génère un fichier Excel"""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl non installé. Exécutez: pip install openpyxl"
            )
        
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        sheets_data = request.sheets if request.sheets else self._get_template_sheets(request)
        
        total_rows = 0
        
        for sheet_data in sheets_data:
            ws = wb.create_sheet(title=sheet_data.name)
            
            # Styles
            header_fill = PatternFill(start_color=self.COLORS["header_bg"], end_color=self.COLORS["header_bg"], fill_type="solid")
            header_font = Font(bold=True, color=self.COLORS["header_text"], size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            for col, header in enumerate(sheet_data.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Données
            for row_idx, row_data in enumerate(sheet_data.rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if isinstance(value, str) and value.startswith("="):
                        cell.value = value
                    else:
                        cell.value = value
                    
                    cell.border = border
                    
                    # Alternance de couleurs
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color=self.COLORS["row_alt"], end_color=self.COLORS["row_alt"], fill_type="solid")
                
                total_rows += 1
            
            # Largeur des colonnes
            for col_idx, header in enumerate(sheet_data.headers, 1):
                col_letter = get_column_letter(col_idx)
                width = sheet_data.column_widths.get(col_letter, len(header) + 5)
                ws.column_dimensions[col_letter].width = width
            
            # Formules personnalisées
            for cell_ref, formula in sheet_data.formulas.items():
                ws[cell_ref] = formula
            
            # Graphiques
            if request.include_charts and sheet_data.charts:
                for chart_config in sheet_data.charts:
                    self._add_chart(ws, chart_config)
        
        # Propriétés du document
        wb.properties.creator = request.author
        wb.properties.title = request.title
        wb.properties.created = datetime.now()
        
        # Sauvegarder
        file_id = f"xlsx_{datetime.now().strftime('%Y%m%d%H%M%S')}_{request.spreadsheet_type.value}"
        filename = f"{file_id}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        wb.save(filepath)
        
        generated = GeneratedSpreadsheet(
            id=file_id,
            filename=filename,
            filepath=filepath,
            spreadsheet_type=request.spreadsheet_type,
            sheet_count=len(sheets_data),
            row_count=total_rows,
            created_at=datetime.now(),
            download_url=f"/skills/xlsx/download/{file_id}"
        )
        
        self.generated_files[file_id] = generated
        
        return generated
    
    def _get_template_sheets(self, request: SpreadsheetRequest) -> List[SheetData]:
        """Génère les feuilles à partir du template"""
        template = self.TEMPLATES.get(request.spreadsheet_type, {"sheets": []})
        
        sheets = []
        for sheet_template in template.get("sheets", []):
            sheets.append(SheetData(
                name=sheet_template["name"],
                headers=sheet_template["headers"],
                rows=sheet_template.get("sample_rows", []),
                formulas=sheet_template.get("summary_formulas", {}),
                column_widths={},
                charts=[]
            ))
        
        return sheets
    
    def _add_chart(self, ws, chart_config: Dict):
        """Ajoute un graphique à la feuille"""
        from openpyxl.chart import BarChart, PieChart, LineChart, Reference
        
        chart_type = chart_config.get("type", "bar")
        title = chart_config.get("title", "Chart")
        data_range = chart_config.get("data_range", "A1:B10")
        position = chart_config.get("position", "H2")
        
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            chart = BarChart()
        
        chart.title = title
        chart.style = 10
        
        ws.add_chart(chart, position)
    
    def get_template(self, spreadsheet_type: SpreadsheetType) -> Dict:
        """Retourne le template pour un type de spreadsheet"""
        return self.TEMPLATES.get(spreadsheet_type, {"sheets": []})
    
    def list_files(self) -> List[GeneratedSpreadsheet]:
        """Liste tous les fichiers générés"""
        return list(self.generated_files.values())


# Instance globale
xlsx_generator = XlsxGenerator()


# Routes API

@router.post("/generate", response_model=Dict[str, Any])
async def generate_spreadsheet(request: SpreadsheetRequest):
    """
    Génère un fichier Excel professionnel
    
    Types supportés:
    - invoice: Facture
    - quote: Devis
    - budget: Budget
    - project: Suivi projet
    - crm: Contacts CRM
    - dashboard: Tableau de bord
    """
    file = await xlsx_generator.generate(request)
    return {
        "status": "success",
        "file_id": file.id,
        "filename": file.filename,
        "sheet_count": file.sheet_count,
        "row_count": file.row_count,
        "download_url": file.download_url
    }


@router.get("/templates/{spreadsheet_type}")
async def get_template(spreadsheet_type: SpreadsheetType):
    """Retourne le template pour un type de spreadsheet"""
    template = xlsx_generator.get_template(spreadsheet_type)
    return {
        "spreadsheet_type": spreadsheet_type.value,
        "template": template
    }


@router.get("/templates")
async def list_templates():
    """Liste tous les templates disponibles"""
    return {
        ss_type.value: template
        for ss_type, template in xlsx_generator.TEMPLATES.items()
    }


@router.get("/files")
async def list_files():
    """Liste tous les fichiers générés"""
    return xlsx_generator.list_files()


@router.get("/download/{file_id}")
async def download_file(file_id: str):
    """Télécharge un fichier généré"""
    from fastapi.responses import FileResponse
    
    file = xlsx_generator.generated_files.get(file_id)
    if not file:
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=file.filepath,
        filename=file.filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.post("/invoice")
async def generate_invoice(
    client_name: str,
    client_company: str,
    items: List[Dict[str, Any]],
    currency: str = "CHF",
    tva_rate: float = 7.7
):
    """
    Génère une facture rapidement
    
    items format: [{"description": "Service", "quantity": 1, "unit_price": 100}]
    """
    rows = []
    for item in items:
        rows.append([
            item.get("description", ""),
            item.get("quantity", 1),
            item.get("unit_price", 0),
            tva_rate,
            f"=B{len(rows)+2}*C{len(rows)+2}",
            f"=E{len(rows)+2}*(1+D{len(rows)+2}/100)"
        ])
    
    request = SpreadsheetRequest(
        spreadsheet_type=SpreadsheetType.INVOICE,
        title=f"Facture - {client_company}",
        sheets=[SheetData(
            name="Facture",
            headers=["Description", "Quantité", f"Prix ({currency})", "TVA %", f"Total HT ({currency})", f"Total TTC ({currency})"],
            rows=rows
        )]
    )
    
    return await generate_spreadsheet(request)
